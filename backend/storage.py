# backend/storage.py
# Модуль для работы с хранилищем данных (Excel/БД)

import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
# Импортируем модель
from .models import TaskRecord

class Storage:
    """Класс для работы с хранилищем данных."""

    def __init__(self, app_config):
        self.config = app_config
        self.filepath = self.config['EXCEL_FILEPATH']
        self.storage_type = self.config['STORAGE_TYPE']

        # Создаем папку data, если её нет
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)

        # Инициализируем Excel файл, если он не существует
        if self.storage_type == 'excel' and not os.path.exists(self.filepath):
            self._initialize_excel_file()

    def _initialize_excel_file(self):
        """Создает новый Excel файл с заголовками."""
        headers = ["Дата", "Время", "День недели", "Часть дня", "Вид задачи", "Задача", "Сложность"]
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        # Настройка ширины колонок (пример)
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        wb.save(self.filepath)
        print(f"Создан новый Excel файл: {self.filepath}")

    def save_task(self, task_record: TaskRecord):
        """Сохраняет одну запись в хранилище."""
        if self.storage_type == 'excel':
            self._save_task_to_excel(task_record)
        elif self.storage_type == 'database':
            # Заглушка для будущей БД
            raise NotImplementedError("Хранилище 'database' ещё не реализовано.")
        else:
            raise ValueError(f"Неизвестный тип хранилища: {self.storage_type}")

    def _save_task_to_excel(self, task_record: TaskRecord):
        """Сохраняет запись в Excel файл."""
        try:
            # Открываем или создаем книгу
            wb = load_workbook(self.filepath) if os.path.exists(self.filepath) else Workbook()
            ws = wb.active

            # Преобразуем сложность в число, если возможно
            try:
                difficulty_value = int(task_record.difficulty)
            except ValueError:
                difficulty_value = task_record.difficulty

            # Убираем переносы строк из описания
            desc_single_line = task_record.description.replace('\n', ' ').replace('\r', ' ')

            # Добавляем строку данных
            ws.append([
                task_record.date,
                task_record.time,
                task_record.weekday,
                task_record.part_of_day,
                task_record.task_type,
                desc_single_line,
                difficulty_value
            ])

            # (Опционально) Обновляем ширину колонок при необходимости
            # for col in ws.columns:
            #     max_len = max((len(str(cell.value)) for cell in col), default=10)
            #     adjusted_width = min(max_len + 2, 50)
            #     if adjusted_width > ws.column_dimensions[col[0].column_letter].width:
            #         ws.column_dimensions[col[0].column_letter].width = adjusted_width

            wb.save(self.filepath)
            print(f"Задача сохранена в Excel: {task_record}")
        except Exception as e:
            print(f"Ошибка при сохранении задачи в Excel: {e}")
            raise # Перебрасываем исключение, чтобы его обработал вызывающий код

    def get_last_tasks(self, count):
        """Получает последние N записей из хранилища."""
        if self.storage_type == 'excel':
            return self._get_last_tasks_from_excel(count)
        elif self.storage_type == 'database':
            # Заглушка для будущей БД
            raise NotImplementedError("Хранилище 'database' ещё не реализовано.")
        else:
            raise ValueError(f"Неизвестный тип хранилища: {self.storage_type}")

    def _get_last_tasks_from_excel(self, count):
        """Получает последние N записей из Excel файла."""
        tasks = []
        try:
            if not os.path.exists(self.filepath):
                 print(f"Excel файл не найден: {self.filepath}")
                 return tasks # Возвращаем пустой список

            wb = load_workbook(self.filepath, read_only=True, data_only=True)
            ws = wb.active

            # Определяем, есть ли заголовок
            has_header = False
            if ws.max_row > 0:
                first_cell_value = ws.cell(row=1, column=1).value
                if isinstance(first_cell_value, str) and 'дата' in first_cell_value.lower():
                    has_header = True

            start_row = 2 if has_header else 1
            total_rows = ws.max_row

            if total_rows >= start_row:
                # Вычисляем индексы строк для чтения
                rows_to_read_start = max(start_row, total_rows - count + 1)
                rows_to_read_end = total_rows

                # Читаем строки с конца
                for row_num in range(rows_to_read_end, rows_to_read_start - 1, -1):
                    row = [cell.value for cell in ws[row_num]]
                    if len(row) >= 7: # Проверка на минимальное количество колонок
                        task = TaskRecord(
                            date=str(row[0]) if row[0] is not None else '',
                            time=str(row[1]) if row[1] is not None else '',
                            weekday=str(row[2]) if row[2] is not None else '',
                            part_of_day=str(row[3]) if row[3] is not None else '',
                            task_type=str(row[4]) if row[4] is not None else 'Р',
                            description=str(row[5]) if row[5] is not None else '',
                            difficulty=str(row[6]) if row[6] is not None else '1'
                        )
                        tasks.append(task)
                        if len(tasks) >= count:
                            break

            # Так как мы читали с конца, переворачиваем список, чтобы последние были в конце
            tasks.reverse()
            wb.close()

        except Exception as e:
            print(f"Ошибка при чтении задач из Excel: {e}")
            # Можно логгировать ошибку или выбрасывать исключение
        return tasks
